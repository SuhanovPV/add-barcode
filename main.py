import openpyxl
import os
import configparser
from reportlab.lib.colors import HexColor
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.graphics import barcode
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.renderPM import drawToFile
from PIL import Image, ImageDraw, ImageFont

CUR_DIR_PATH = os.path.dirname(os.path.realpath(__file__))
CONFIG_FILE = os.path.join(CUR_DIR_PATH, "config.ini")
BARCODE_FILE = os.path.join(CUR_DIR_PATH, "_barcode.jpg")
RESULT_DIR = os.path.join(CUR_DIR_PATH, "results")

config = configparser.ConfigParser()
config.read(CONFIG_FILE)

BC_WIDTH = int(config["BARCODE"]["width"])
BC_HEIGHT = int(config["BARCODE"]["height"])
BC_x = int(config["BARCODE"]["x"])
BC_y = int(config["BARCODE"]["y"])
BC_border_v = int(config["BARCODE"]["border_v"])
BC_border_h = int(config["BARCODE"]["border_h"])
BC_COLOR = config["BARCODE"]["color"]
BC_TEXT_COLOR = config["BARCODE"]["text_color"]
BC_RATIO = 1.4

TEXT_x = int(config["TEXT"]["text_x"])
TEXT_y = int(config["TEXT"]["text_y"])
TEXT_SIZE = int(config["TEXT"]["font_size"])
TEXT_COLOR = config["TEXT"]["font_color"]

EXCEL_EXT = [".xsl", ".xlsx", ".XSL", ".XLSX"]
PICTURE_EXT = [".jpg", ".jpeg", ".bmp", ".png", ".JPG", ".JPEG", ".BMP", ".PNG"]
FONT_DIR = "fonts"
FONT_OTF = "ALS_Granate_Book_1.1.otf"
FONT_TTF = "ALS_Granate_Book_1.1.ttf"




def create_barcode(code):
    """Save tmp barcode file with 'code' data"""
    width = BC_WIDTH
    height = int(width / BC_RATIO)
    draw = Drawing(width, height)
    pdfmetrics.registerFont(TTFont('Granate_Book', os.path.abspath(os.path.join(FONT_DIR, FONT_TTF))))
    new_barcode = barcode.createBarcodeDrawing('EAN13', value=code, height=height, width=width,
                                               barFillColor=HexColor(BC_COLOR), fontName='Granate_Book',
                                               textColor=HexColor(BC_TEXT_COLOR))
    draw.add(new_barcode)
    drawToFile(draw, BARCODE_FILE)


def is_need_crop():
    return BC_WIDTH / BC_RATIO > BC_HEIGHT


def put_barcode_to_cert(image):
    bc = Image.open(BARCODE_FILE)
    if is_need_crop():
        image.paste(bc.crop((0, bc.height - BC_HEIGHT, bc.width, bc.height)), (BC_x, BC_y))
    else:
        image.paste(bc, (BC_x, BC_y))
    bc.close()
    return image


def put_bc_background(image):
    draw = ImageDraw.Draw(image)
    draw.rectangle((BC_x - BC_border_h, BC_y, BC_x + BC_WIDTH, BC_y + BC_HEIGHT + BC_border_v), fill="#FFFFFF")
    return image


def put_text_to_cert(image, text):
    draw = ImageDraw.Draw(image)
    font = os.path.abspath(os.path.join(FONT_DIR, FONT_OTF))
    text_font = ImageFont.truetype(font=font, size=TEXT_SIZE)
    draw.text((TEXT_x, TEXT_y), text, font=text_font, fill=TEXT_COLOR)
    return image


def insert_data_to_picture(cert_filename, code, price):
    print(f"create barcode {code}")
    create_barcode(code)
    cert = Image.open(cert_filename).convert('RGB')
    put_bc_background(cert)
    cert = put_barcode_to_cert(cert)
    cert = put_text_to_cert(cert, f"{price} ₽")
    cert.save(f"{RESULT_DIR}/{code}.jpg",
              format="JPEG",
              quality=100,
              icc_profile=cert.info.get('icc_profile', ''))
    cert.close()
    os.remove(BARCODE_FILE)


def get_data_from_xsl(file_name):
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook.active

    for i in range(0, worksheet.max_row):
        data = []
        for col in worksheet.iter_cols(1, 2):
            if col[i].value is not None:
                data.append(col[i].value)
        if len(data) == 2:
            yield data


def create_dir(name):
    if not os.path.exists(name):
        os.mkdir(name)


def get_filename(extension):
    files_list = [f for f in os.listdir(CUR_DIR_PATH) if os.path.isfile(f) and os.path.splitext(f)[1] in extension]
    if files_list:
        return files_list[0]
    print(f"Не найдено файла с расширением {extension}")
    return None


def remove_tmp_files():
    """remove accessory files"""
    tmp_files = [f for f in os.listdir(CUR_DIR_PATH) if os.path.isfile(f) and os.path.splitext(f)[1] in PICTURE_EXT]
    for f in tmp_files:
        if f.startswith("_"):
            os.remove(f)


def init():
    create_dir(RESULT_DIR)
    remove_tmp_files()


if __name__ == "__main__":
    init()
    create_dir(RESULT_DIR)
    excel_filename = get_filename(EXCEL_EXT)
    cert_filename = get_filename(PICTURE_EXT)
    for data in get_data_from_xsl(excel_filename):
        insert_data_to_picture(cert_filename, *data)
    print("DONE! Check result folder.\nfor exit press eny key >", end=" ")
    input()
